import xml.etree.ElementTree as ET
import requests
from glob import glob
import logging
from openpyxl import load_workbook
from typing import List
import re
from pprint import pprint

from .db import db
from . import dbmodel

BASE_URL="https://fachportal.gematik.de/fachportal-import/files/"

FORMAT = '[convert] %(levelname)s: %(message)s'
logging.basicConfig(format=FORMAT, level=logging.WARNING)


def convert_data():
    index_validity = ET.parse("./raw-data/metadaten-xml/index-validity-status-from-excel-file.xml")

    indices = glob("./raw-data/metadaten-xml/index-characteristics-*.xml")
    indices.sort(reverse=True)

    for index in indices:
        xml = ET.parse(index)
        index = index.rsplit('/', 1)[-1]

        for file_el in xml.getroot().iter('file'):
            id = file_el.find('id').text

            # this version breaks the import
            if id.startswith('gemProdT_Kon_PTV5_ALT'):
                continue

            if id.startswith('gemProdT_Kon_PTV'):
                id = 'gemProdT_Kon'


            version = file_el.find('typeVersion').text
            descver = file_el.find('docVersion').text

            filenames = file_el.find('fileNames').findall('fileName')

            url = requests.compat.urljoin(BASE_URL, f"../steckbriefe/{filenames[0].text}")
            r = requests.head(url)
            if r.status_code != 200:
                error = f"URL not found {url}, referenced in {id}, v{version}, {index}"
                logging.error (error)

            validity_el = index_validity.findall(f".//file[id='{id}']")


            if len(validity_el) == 0:
                logging.warning(f"Unknown entity {id}, v{version} from {index}")
            else:
                pass

            if id.startswith('gemProdT'):
                type = dbmodel.SubjectType.Product
            elif id.startswith('gemAnbT'):
                type = dbmodel.SubjectType.Provider
            elif id.startswith('gemAnw'):
                type = dbmodel.SubjectType.ThirdParty
            elif id.startswith('gemVZ'):
                type = dbmodel.SubjectType.Requirements
            else:
                logging.error(f"Unknown type: {id}")
                continue

            title = file_el.find('description').text
            title = re.sub(r'steckbrief', '', title)
            title = re.sub(r' Prüfvorschrift', '', title)

            subj_vers = dbmodel.SubjectVersion(subject_id=id, title=title, type=type, version=version)
            db.SubjectVersion.update({'subject_id':subj_vers.subject_id, 'version':subj_vers.version}, subj_vers.dict(), upsert=True)
            logging.info(f"Subject {id}, v{version}")

            # read requirements for this subject
            xlsx_filename = filenames[1].text
            xlsx_files = glob(f"./raw-data/**/{xlsx_filename}", recursive=True)

            if len(xlsx_files) == 0:
                logging.error(f"XLSX File not found for {xlsx_filename} in {index}")
                subj_reqs = dict()
            else:
                subj_reqs = load_subject_version_requirements(xlsx_files[0])

            for doc_el in file_el.findall('docs/doc'):
                rid = doc_el.find('id').text
                rver = doc_el.find('docVersion').text

                resource = dbmodel.Resource(id=rid, description=doc_el.find('description').text)
                db.Resource.update({'id':rid}, resource.dict(), upsert=True)

                rv_dict = db.ResourceVersion.find_one({'resource_id': resource.id, 'version': rver})

                if rv_dict != None:
                    logging.debug(f"Already exists: {resource.id}, v{rver}")
                    resource_version = dbmodel.ResourceVersion(**rv_dict)
                else:
                    url = requests.compat.urljoin(BASE_URL, f"{rid}_V{rver}.pdf")
                    r = requests.head(url)
                    if r.status_code != 200:
                        error = f"URL not found for document {rid}_V{rver}.pdf, referenced in {subj_vers.subject_id}, v{subj_vers.version}, {index}"
                        logging.error (error)
                        url = error

                    resource_version = dbmodel.ResourceVersion(resource_id= resource.id, version=rver, url=url)
                    db.ResourceVersion.update({'resource_id':resource.id, 'version':rver}, resource_version.dict(), upsert=True)
                    db.ResourceVersion.find_one({'resource_id': resource.id, 'version': rver})
                    logging.info(f"ResourceVersion {resource_version.resource_id}, v{resource_version.version}")
              
                resource_ref = dbmodel.ResourceReference(resource_id=resource_version.resource_id, resource_version=resource_version.version)
                resource_ref.requirements.extend(subj_reqs.get(resource_version.resource_id, list()))
                subj_vers.references.append(resource_ref) 

            db.SubjectVersion.update({'subject_id':subj_vers.subject_id, 'version':subj_vers.version}, subj_vers.dict(), upsert=True)
    
        # uncomment for fast partial import
        #break
                
#pprint (load_requirements(doc_filename, 'Anforderungen'))
#pprint (load_requirements(desc_filename, 'Blattanforderungen'))
def load_subject_version_requirements(filename: str):
    result = dict()
    wb = load_workbook(filename)
    ws = wb['Blattanforderungen']
    for row in ws.iter_rows(min_row=2):
        req = dbmodel.RequirementReference(
            id=row[0].value,
            title=row[1].value,
            text=row[2].value,
            html=row[3].value,
            level=row[4].value,
            test_procedure=row[6].value,
        )
        resource_id = row[5].value
        if not resource_id in result:
            result[resource_id] = list()
        result.get(resource_id).append(req)

    return result
